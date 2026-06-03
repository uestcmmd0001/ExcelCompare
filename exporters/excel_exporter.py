from __future__ import annotations

"""Excel report exporter."""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from compare_excel_to_excel.models.schemas import DiffItem
from compare_excel_to_excel.planning.plan_schema import to_dict
from compare_excel_to_excel.processing.differ import diff_summary


HEADERS = [
    "序号",
    "结果分层",
    "差异类型",
    "严重程度",
    "Sheet",
    "业务主键",
    "原行号",
    "新行号",
    "列名",
    "原值",
    "新值",
    "说明",
    "是否需人工复核",
    "规则来源",
    "diff_id",
]

SEVERITY_FILL = {
    "high": "F4CCCC",
    "medium": "FFF2CC",
    "low": "D9EAD3",
}

LAYER_FILL = {
    "structure_change": "CFE2F3",
    "match_risk": "D9D9D9",
}


def export_excel(
    run_dir: str | Path,
    diff_items: list[DiffItem],
    *,
    comparison_plan: Any | None = None,
    plan_validation: Any | None = None,
) -> Path:
    output_path = Path(run_dir) / "output" / "diff_results.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "汇总"
    _write_summary(summary_ws, diff_items, plan_validation)

    if comparison_plan is not None or plan_validation is not None:
        _write_plan_sheet(wb.create_sheet("比对计划说明"), comparison_plan, plan_validation)
        _write_plan_comparison_sheet(wb.create_sheet("计划对比"), comparison_plan, plan_validation)
        _write_plan_risk_sheet(wb.create_sheet("计划风险"), plan_validation)

    sheet_defs = [
        ("确认差异", lambda d: d.result_layer == "confirmed_diff"),
        ("结构变化", lambda d: d.result_layer == "structure_change"),
        ("匹配风险", lambda d: d.result_layer == "match_risk"),
        ("需人工复核", lambda d: d.result_layer == "review_item"),
        ("新增行", lambda d: d.diff_type == "row_added"),
        ("删除行", lambda d: d.diff_type == "row_deleted"),
        ("修改单元格", lambda d: d.diff_type in ("cell_modified", "formula_modified")),
        ("全部结果", lambda d: True),
    ]

    for title, predicate in sheet_defs:
        _write_diff_sheet(wb.create_sheet(title), [item for item in diff_items if predicate(item)])

    wb.save(output_path)
    return output_path


def _write_summary(ws, diff_items: list[DiffItem], plan_validation: Any | None = None) -> None:
    summary = diff_summary(diff_items)
    validation = to_dict(plan_validation) if plan_validation is not None else {}
    rows = [
        ["指标", "数量"],
        ["差异总数", summary["total"]],
        ["确认差异", summary["by_layer"].get("confirmed_diff", 0)],
        ["结构变化", summary["by_layer"].get("structure_change", 0)],
        ["匹配风险", summary["by_layer"].get("match_risk", 0)],
        ["需人工复核", summary["by_layer"].get("review_item", 0)],
        ["高风险", summary["by_severity"].get("high", 0)],
        ["中风险", summary["by_severity"].get("medium", 0)],
        ["低风险", summary["by_severity"].get("low", 0)],
        ["比对计划整体置信度", validation.get("overall_confidence", "")],
        ["计划风险数量", len(validation.get("risks", []) or [])],
    ]

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    start = len(rows) + 3
    ws.cell(row=start, column=1, value="差异类型")
    ws.cell(row=start, column=2, value="数量")
    ws.cell(row=start, column=1).font = Font(bold=True)
    ws.cell(row=start, column=2).font = Font(bold=True)
    for offset, (diff_type, count) in enumerate(summary["by_type"].items(), start=1):
        ws.cell(row=start + offset, column=1, value=diff_type)
        ws.cell(row=start + offset, column=2, value=count)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "A2"


def _write_plan_sheet(ws, comparison_plan: Any | None, plan_validation: Any | None) -> None:
    plan = to_dict(comparison_plan) if comparison_plan is not None else {}
    validation = to_dict(plan_validation) if plan_validation is not None else {}
    primary = plan.get("primary_plan", {}) or {}
    sheet_validations = validation.get("sheet_validations", []) or []
    validations_by_pair = {
        (item.get("baseline_sheet"), item.get("revised_sheet")): item
        for item in sheet_validations
    }

    ws.append(["字段", "内容"])
    ws.append(["规划器", plan.get("planner", "")])
    ws.append(["计划来源", plan.get("source", "")])
    ws.append(["主计划策略", primary.get("strategy", "")])
    ws.append(["主计划置信度", primary.get("confidence", "")])
    ws.append(["校验后整体置信度", validation.get("overall_confidence", "")])
    ws.append([])

    headers = [
        "baseline sheet",
        "revised sheet",
        "表头行 baseline",
        "表头行 revised",
        "采用主键",
        "只作结构留痕列",
        "完全忽略列",
        "数字列",
        "日期列",
        "sheet匹配分",
        "表头分",
        "主键覆盖分",
        "主键唯一分",
        "列映射分",
        "类型一致分",
        "综合置信度",
        "依据",
        "不确定点/风险",
    ]
    start = ws.max_row + 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, sheet_plan in enumerate(primary.get("sheet_pairs", []) or [], start=start + 1):
        validation_item = validations_by_pair.get(
            (sheet_plan.get("baseline_sheet"), sheet_plan.get("revised_sheet")),
            {},
        )
        row = [
            sheet_plan.get("baseline_sheet", ""),
            sheet_plan.get("revised_sheet", ""),
            sheet_plan.get("header_row_baseline", ""),
            sheet_plan.get("header_row_revised", ""),
            ", ".join(validation_item.get("selected_key_columns") or sheet_plan.get("key_columns") or []),
            ", ".join(sheet_plan.get("structure_only_columns") or []),
            ", ".join(sheet_plan.get("ignore_columns") or []),
            _format_mapping(sheet_plan.get("numeric_columns") or {}),
            ", ".join(sheet_plan.get("date_columns") or []),
            validation_item.get("sheet_match_score", ""),
            validation_item.get("header_score", ""),
            validation_item.get("key_coverage_score", ""),
            validation_item.get("key_uniqueness_score", ""),
            validation_item.get("column_mapping_score", ""),
            validation_item.get("type_consistency_score", ""),
            validation_item.get("overall_confidence", ""),
            "；".join(sheet_plan.get("reasons") or []),
            "；".join((sheet_plan.get("uncertainties") or []) + (validation_item.get("risks") or [])),
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [22, 22, 12, 12, 28, 30, 24, 24, 24, 12, 12, 12, 12, 12, 12, 12, 42, 52]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = f"A{start + 1}"
    ws.auto_filter.ref = ws.dimensions


def _write_plan_comparison_sheet(ws, comparison_plan: Any | None, plan_validation: Any | None) -> None:
    plan = to_dict(comparison_plan) if comparison_plan is not None else {}
    validation = to_dict(plan_validation) if plan_validation is not None else {}
    primary = plan.get("primary_plan", {}) or {}
    alternatives = plan.get("alternative_plans", []) or []

    headers = [
        "计划角色",
        "计划ID",
        "规划器",
        "策略",
        "计划置信度",
        "Sheet数量",
        "Sheet对",
        "主键摘要",
        "结构留痕列摘要",
        "完全忽略列摘要",
        "校验后整体置信度",
        "说明/不确定点",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = [("primary", primary)]
    rows.extend((f"alternative_{idx}", item) for idx, item in enumerate(alternatives, start=1))

    for row_idx, (role, candidate) in enumerate(rows, start=2):
        sheet_pairs = candidate.get("sheet_pairs", []) or []
        row = [
            role,
            candidate.get("candidate_id", ""),
            plan.get("planner", ""),
            candidate.get("strategy", ""),
            candidate.get("confidence", ""),
            len(sheet_pairs),
            _sheet_pair_summary(sheet_pairs),
            _sheet_pair_field_summary(sheet_pairs, "key_columns"),
            _sheet_pair_field_summary(sheet_pairs, "structure_only_columns"),
            _sheet_pair_field_summary(sheet_pairs, "ignore_columns"),
            validation.get("overall_confidence", "") if role == "primary" else "",
            "；".join((candidate.get("reasons") or []) + (candidate.get("uncertainties") or [])),
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if not rows:
        ws.cell(row=2, column=1, value="未生成可对比计划")

    widths = [14, 22, 20, 30, 12, 10, 34, 34, 34, 28, 16, 70]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_plan_risk_sheet(ws, plan_validation: Any | None) -> None:
    validation = to_dict(plan_validation) if plan_validation is not None else {}
    headers = ["序号", "风险/说明"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    risks = validation.get("risks", []) or []
    if not risks:
        risks = ["当前自动比对计划未发现阻断性风险；仍建议结合业务语境复核高风险差异。"]

    for row_idx, risk in enumerate(risks, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=risk)
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"


def _write_diff_sheet(ws, diff_items: list[DiffItem]) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for row_idx, item in enumerate(diff_items, start=2):
        row_data = [
            row_idx - 1,
            _layer_label(item.result_layer),
            _type_label(item.diff_type),
            _severity_label(item.severity),
            item.sheet_name,
            item.business_key or "",
            item.row_before or "",
            item.row_after or "",
            item.column_name or "",
            item.old_value,
            item.new_value,
            item.summary,
            "是" if item.need_human_review else "否",
            item.rule_source,
            item.diff_id,
        ]
        fill_color = _fill_color(item)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill

    widths = [8, 14, 16, 10, 20, 26, 10, 10, 18, 32, 32, 46, 14, 18, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _layer_label(layer: str) -> str:
    return {
        "confirmed_diff": "确认差异",
        "structure_change": "结构变化",
        "match_risk": "匹配风险",
        "review_item": "需人工复核",
    }.get(layer, layer)


def _type_label(diff_type: str) -> str:
    return {
        "sheet_added": "Sheet新增",
        "sheet_deleted": "Sheet删除",
        "column_added": "列新增",
        "column_deleted": "列删除",
        "row_added": "行新增",
        "row_deleted": "行删除",
        "row_moved": "行移动",
        "cell_modified": "单元格修改",
        "formula_modified": "公式修改",
        "duplicate_key": "主键重复",
        "duplicate_key_rows_skipped": "重复主键跳过逐格比对",
    }.get(diff_type, diff_type)


def _severity_label(severity: str) -> str:
    return {"high": "高", "medium": "中", "low": "低"}.get(severity, severity)


def _format_mapping(mapping: dict[str, Any]) -> str:
    return "；".join(f"{key}:{value}" for key, value in mapping.items())


def _sheet_pair_summary(sheet_pairs: list[dict[str, Any]]) -> str:
    parts = []
    for item in sheet_pairs:
        parts.append(f"{item.get('baseline_sheet', '')} -> {item.get('revised_sheet', '')}")
    return "；".join(parts)


def _sheet_pair_field_summary(sheet_pairs: list[dict[str, Any]], field_name: str) -> str:
    parts = []
    for item in sheet_pairs:
        values = item.get(field_name) or []
        if values:
            parts.append(f"{item.get('baseline_sheet', '')}: {', '.join(values)}")
    return "；".join(parts)


def _fill_color(item: DiffItem) -> str:
    return LAYER_FILL.get(item.result_layer, SEVERITY_FILL.get(item.severity, "FFFFFF"))

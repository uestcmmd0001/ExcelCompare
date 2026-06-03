from __future__ import annotations

"""Parse .xlsx files into structured workbook models."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from compare_excel_to_excel.config import get_sheet_rule, resolve_auto_rule, should_compare_sheet
from compare_excel_to_excel.models.schemas import ExcelCell, ExcelRow, ExcelSheet, ExcelWorkbook
from compare_excel_to_excel.processing.normalizer import normalize_value


def parse_workbook(path: str | Path, config: dict[str, Any], role: str | None = None) -> ExcelWorkbook:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() == ".xls":
        return _parse_xls_workbook(workbook_path, config, role)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx/.xls is supported: {workbook_path}")

    wb = load_workbook(str(workbook_path), data_only=False)
    data_wb = load_workbook(str(workbook_path), data_only=True)
    sheets: dict[str, ExcelSheet] = {}

    for ws in wb.worksheets:
        if not should_compare_sheet(config, ws.title):
            continue
        rule = get_sheet_rule(config, ws.title, role)
        rule = resolve_auto_rule(rule, _xlsx_preview_rows(ws), config)
        data_ws = data_wb[ws.title]
        sheets[ws.title] = _parse_sheet(ws, data_ws, rule)

    return ExcelWorkbook(path=str(workbook_path), sheets=sheets)


def _parse_xls_workbook(workbook_path: Path, config: dict[str, Any], role: str | None = None) -> ExcelWorkbook:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency failure path
        raise RuntimeError("xlrd is required to parse .xls files") from exc

    wb = xlrd.open_workbook(str(workbook_path), formatting_info=True)
    sheets: dict[str, ExcelSheet] = {}
    for ws in wb.sheets():
        if not should_compare_sheet(config, ws.name):
            continue
        rule = get_sheet_rule(config, ws.name, role)
        rule = resolve_auto_rule(rule, _xls_preview_rows(ws), config)
        sheets[ws.name] = _parse_xls_sheet(ws, rule)

    return ExcelWorkbook(path=str(workbook_path), sheets=sheets)


def _parse_sheet(ws, data_ws, rule) -> ExcelSheet:
    header_cells = list(ws[rule.header_row])
    columns: list[str] = []
    column_indexes: list[int] = []
    seen: dict[str, int] = {}

    for cell in header_cells:
        if _is_blank(cell.value) and _is_empty_column(ws, cell.column, rule.header_row + 1):
            continue
        name = _header_name(cell.value, cell.column)
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)
        column_indexes.append(cell.column)

    rows: list[ExcelRow] = []
    for row_idx in range(rule.header_row + 1, ws.max_row + 1):
        cells: dict[str, ExcelCell] = {}
        has_value = False
        for column_name, col_idx in zip(columns, column_indexes):
            cell = ws.cell(row=row_idx, column=col_idx)
            raw_value = cell.value
            formula = raw_value if isinstance(raw_value, str) and raw_value.startswith("=") else None
            cached_value = data_ws.cell(row=row_idx, column=col_idx).value if formula else None
            value = cached_value if cached_value is not None else raw_value
            normalized = normalize_value(value, column_name, rule)
            if normalized:
                has_value = True
            cells[column_name] = ExcelCell(
                column_name=column_name,
                coordinate=cell.coordinate,
                value=value,
                formula=formula,
                normalized_value=normalized,
            )

        if not has_value:
            continue

        key, key_source = _build_row_key(cells, row_idx, rule)
        rows.append(ExcelRow(row_number=row_idx, key=key, key_source=key_source, cells=cells))

    hidden_columns = [
        get_column_letter(idx)
        for idx in range(1, ws.max_column + 1)
        if ws.column_dimensions[get_column_letter(idx)].hidden
    ]
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]

    return ExcelSheet(
        name=ws.title,
        header_row=rule.header_row,
        columns=columns,
        rows=rows,
        hidden_columns=hidden_columns,
        merged_ranges=merged_ranges,
    )


def _parse_xls_sheet(ws, rule) -> ExcelSheet:
    header_idx = rule.header_row - 1
    if header_idx < 0 or header_idx >= ws.nrows:
        return ExcelSheet(name=ws.name, header_row=rule.header_row, columns=[], rows=[])

    columns: list[str] = []
    column_indexes: list[int] = []
    seen: dict[str, int] = {}
    for col_idx in range(ws.ncols):
        header_value = _xls_cell_value(ws, header_idx, col_idx)
        if _is_blank(header_value) and _is_empty_xls_column(ws, col_idx, header_idx + 1):
            continue
        name = _header_name(header_value, col_idx + 1)
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)
        column_indexes.append(col_idx)

    rows: list[ExcelRow] = []
    for row_idx in range(header_idx + 1, ws.nrows):
        cells: dict[str, ExcelCell] = {}
        has_value = False
        for column_name, col_idx in zip(columns, column_indexes):
            value = _xls_cell_value(ws, row_idx, col_idx)
            normalized = normalize_value(value, column_name, rule)
            if normalized:
                has_value = True
            coordinate = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
            cells[column_name] = ExcelCell(
                column_name=column_name,
                coordinate=coordinate,
                value=value,
                formula=None,
                normalized_value=normalized,
            )

        if not has_value:
            continue

        key, key_source = _build_row_key(cells, row_idx + 1, rule)
        rows.append(ExcelRow(row_number=row_idx + 1, key=key, key_source=key_source, cells=cells))

    hidden_columns = []
    for col_idx, colinfo in ws.colinfo_map.items():
        if getattr(colinfo, "hidden", False):
            hidden_columns.append(get_column_letter(col_idx + 1))

    merged_ranges = [
        f"{get_column_letter(c_low + 1)}{r_low + 1}:{get_column_letter(c_high)}{r_high}"
        for r_low, r_high, c_low, c_high in ws.merged_cells
    ]

    return ExcelSheet(
        name=ws.name,
        header_row=rule.header_row,
        columns=columns,
        rows=rows,
        hidden_columns=hidden_columns,
        merged_ranges=merged_ranges,
    )


def _header_name(value: Any, column: int) -> str:
    if value is None:
        return f"__EMPTY_COL_{get_column_letter(column)}"
    return str(value).strip() or f"__EMPTY_COL_{get_column_letter(column)}"


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_empty_column(ws, column: int, start_row: int) -> bool:
    for row_idx in range(start_row, ws.max_row + 1):
        if ws.cell(row=row_idx, column=column).value not in (None, ""):
            return False
    return True


def _is_empty_xls_column(ws, column: int, start_row: int) -> bool:
    for row_idx in range(start_row, ws.nrows):
        if _xls_cell_value(ws, row_idx, column) not in (None, ""):
            return False
    return True


def _xls_cell_value(ws, row_idx: int, col_idx: int) -> Any:
    cell = ws.cell(row_idx, col_idx)
    if cell.ctype == 0:
        return None
    if cell.ctype == 2 and float(cell.value).is_integer():
        return int(cell.value)
    if cell.ctype == 3:
        try:
            date_tuple = ws.book.datemode
            import xlrd

            return xlrd.xldate_as_datetime(cell.value, date_tuple)
        except Exception:
            return cell.value
    if cell.ctype == 4:
        return bool(cell.value)
    return cell.value


def _build_row_key(cells: dict[str, ExcelCell], row_idx: int, rule) -> tuple[str, str]:
    if rule.key_columns and all(col in cells for col in rule.key_columns):
        parts = [cells[col].normalized_value for col in rule.key_columns]
        if any(parts):
            return "|".join(parts), "business_key"
    return f"__row_{row_idx}", "row_number"


def _xlsx_preview_rows(ws, max_rows: int = 80) -> list[list[Any]]:
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_rows), values_only=True):
        rows.append(list(row))
    return rows


def _xls_preview_rows(ws, max_rows: int = 80) -> list[list[Any]]:
    rows = []
    for row_idx in range(min(ws.nrows, max_rows)):
        rows.append([_xls_cell_value(ws, row_idx, col_idx) for col_idx in range(ws.ncols)])
    return rows

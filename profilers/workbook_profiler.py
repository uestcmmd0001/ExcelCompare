from __future__ import annotations

"""Infer sheet structure and suggest comparison config for unknown Excel files."""

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def profile_workbook_pair(
    baseline_path: str | Path,
    revised_path: str | Path,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = config or {}
    baseline = profile_workbook(baseline_path, config)
    revised = profile_workbook(revised_path, config)
    suggested_config = suggest_pair_config(baseline, revised, config)
    return {
        "baseline": baseline,
        "revised": revised,
        "suggested_config": suggested_config,
    }


def profile_workbook(path: str | Path, config: dict[str, Any] | None = None) -> dict[str, Any]:
    config = config or {}
    workbook_path = Path(path)
    suffix = workbook_path.suffix.lower()
    if suffix == ".xls":
        sheets = _profile_xls(workbook_path, config)
    elif suffix == ".xlsx":
        sheets = _profile_xlsx(workbook_path, config)
    else:
        raise ValueError(f"Only .xlsx/.xls is supported: {workbook_path}")

    return {
        "path": str(workbook_path),
        "sheets": sheets,
    }


def suggest_pair_config(
    baseline_profile: dict[str, Any],
    revised_profile: dict[str, Any],
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = config or {}
    base_sheets = {s["name"]: s for s in baseline_profile.get("sheets", [])}
    rev_sheets = {s["name"]: s for s in revised_profile.get("sheets", [])}
    common_sheet_names = [name for name in base_sheets if name in rev_sheets]

    sheet_configs: dict[str, Any] = {}
    include_sheets: list[str] = []
    for sheet_name in common_sheet_names:
        base_sheet = base_sheets[sheet_name]
        rev_sheet = rev_sheets[sheet_name]
        if base_sheet.get("data_rows", 0) == 0 and rev_sheet.get("data_rows", 0) == 0:
            continue
        if not _should_auto_include_sheet(base_sheet, rev_sheet, config):
            continue

        include_sheets.append(sheet_name)
        header_row = _choose_header_row(base_sheet, rev_sheet)
        common_columns = [
            col for col in base_sheet.get("columns", [])
            if col in set(rev_sheet.get("columns", [])) and not _is_generated_empty_column(col)
        ]
        key_columns = _suggest_key_columns(base_sheet, rev_sheet, common_columns, config)
        structure_only_columns = _suggest_structure_only_columns(base_sheet, rev_sheet)

        sheet_configs[sheet_name] = {
            "header_row": header_row,
            "key_columns": key_columns,
            "ignore_columns": [],
            "structure_only_columns": structure_only_columns,
            "numeric_columns": _suggest_numeric_columns(base_sheet, rev_sheet, config),
            "date_columns": _suggest_date_columns(base_sheet, rev_sheet, config),
            "compare_formulas": True,
        }

    return {
        "default": {
            "header_row": "auto",
            "key_columns": "auto",
            "ignore_columns": [],
            "structure_only_columns": [],
            "numeric_columns": {},
            "date_columns": [],
            "case_insensitive_columns": [],
            "compare_formulas": True,
            "blank_values": ["", "NULL", "N/A", "NA", "-"],
        },
        "include_sheets": include_sheets,
        "sheets": sheet_configs,
    }


def _profile_xlsx(path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    result = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 1, int(config.get("profile_max_rows", 80))),
            values_only=True,
        ):
            rows.append(list(row))
        result.append(_profile_rows(ws.title, rows, ws.max_row or 0, ws.max_column or 0, config))
    return result


def _profile_xls(path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("xlrd is required to profile .xls files") from exc

    wb = xlrd.open_workbook(str(path), formatting_info=False)
    result = []
    max_rows = int(config.get("profile_max_rows", 80))
    for ws in wb.sheets():
        rows = []
        for row_idx in range(min(ws.nrows, max_rows)):
            rows.append([ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)])
        result.append(_profile_rows(ws.name, rows, ws.nrows, ws.ncols, config))
    return result


def _profile_rows(sheet_name: str, rows: list[list[Any]], total_rows: int, total_cols: int, config: dict[str, Any]) -> dict[str, Any]:
    header_row = infer_header_row(rows, config)
    columns = _columns_from_row(rows[header_row - 1] if header_row and header_row <= len(rows) else [])
    data_rows = max(total_rows - header_row, 0) if header_row else 0
    column_profiles = _profile_columns(rows, header_row, columns)
    key_candidates = _rank_key_candidates(column_profiles, data_rows, config)
    return {
        "name": sheet_name,
        "total_rows": total_rows,
        "total_columns": total_cols,
        "detected_header_row": header_row,
        "columns": columns,
        "data_rows": data_rows,
        "duplicate_header_ratio": _duplicate_header_ratio(columns),
        "key_candidates": key_candidates[:6],
        "column_profiles": column_profiles,
        "row_samples": _row_samples(rows, header_row, columns, config),
    }


def infer_header_row(rows: list[list[Any]], config: dict[str, Any] | None = None) -> int:
    config = config or {}
    best_row = 1
    best_score = -1.0
    for idx, row in enumerate(rows[:20], start=1):
        values = [_clean(v) for v in row]
        non_blank = [v for v in values if v]
        if len(non_blank) < 2:
            continue

        unique_ratio = len(set(non_blank)) / max(len(non_blank), 1)
        text_ratio = sum(1 for v in non_blank if not _looks_numeric(v)) / max(len(non_blank), 1)
        below_non_blank = 0
        for below in rows[idx: min(idx + 8, len(rows))]:
            below_non_blank += sum(1 for v in below if _clean(v))

        keyword_hits = sum(1 for v in non_blank if _keyword_score(v, config) > 0)
        score = len(non_blank) * 1.5 + unique_ratio * 5 + text_ratio * 3 + keyword_hits * 2 + min(below_non_blank, 20) * 0.2
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def _columns_from_row(row: list[Any]) -> list[str]:
    columns = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(row, start=1):
        name = _clean(value)
        if not name:
            continue
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)
    return columns


def _profile_columns(rows: list[list[Any]], header_row: int, columns: list[str]) -> list[dict[str, Any]]:
    if not header_row or not columns:
        return []
    header_values = [_clean(v) for v in rows[header_row - 1]]
    col_indexes = [idx for idx, value in enumerate(header_values) if value][:len(columns)]
    data_rows = rows[header_row: min(header_row + 200, len(rows))]
    profiles = []
    for name, col_idx in zip(columns, col_indexes):
        values = [_clean(row[col_idx]) if col_idx < len(row) else "" for row in data_rows]
        non_blank = [v for v in values if v]
        unique_count = len(set(non_blank))
        duplicate_count = len(non_blank) - unique_count
        profiles.append({
            "name": name,
            "index": col_idx,
            "non_blank_count": len(non_blank),
            "unique_count": unique_count,
            "duplicate_count": duplicate_count,
            "unique_ratio": round(unique_count / max(len(non_blank), 1), 4),
            "numeric_ratio": round(sum(1 for v in non_blank if _looks_numeric(v)) / max(len(non_blank), 1), 4),
            "date_like_ratio": round(sum(1 for v in non_blank if _looks_date_like(v)) / max(len(non_blank), 1), 4),
        })
    return profiles


def _row_samples(
    rows: list[list[Any]],
    header_row: int,
    columns: list[str],
    config: dict[str, Any],
) -> list[dict[str, str]]:
    if not header_row or not columns:
        return []
    header_values = [_clean(v) for v in rows[header_row - 1]]
    col_indexes = [idx for idx, value in enumerate(header_values) if value][:len(columns)]
    max_rows = int(config.get("profile_key_sample_rows", 200))
    samples = []
    for row in rows[header_row: min(header_row + max_rows, len(rows))]:
        sample: dict[str, str] = {}
        has_value = False
        for name, col_idx in zip(columns, col_indexes):
            value = _clean(row[col_idx]) if col_idx < len(row) else ""
            sample[name] = value
            has_value = has_value or bool(value)
        if has_value:
            samples.append(sample)
    return samples


def _rank_key_candidates(column_profiles: list[dict[str, Any]], data_rows: int, config: dict[str, Any]) -> list[dict[str, Any]]:
    candidates = []
    for profile in column_profiles:
        if profile["non_blank_count"] == 0:
            continue
        coverage = profile["non_blank_count"] / max(data_rows, 1)
        uniqueness = profile["unique_ratio"]
        score = coverage * 30 + uniqueness * 50 + _keyword_score(profile["name"], config)
        if coverage < 0.5:
            score -= 20
        if profile["numeric_ratio"] > 0.95 and _keyword_score(profile["name"], config) < 6:
            score -= 35
        candidates.append({
            "columns": [profile["name"]],
            "score": round(score, 2),
            "coverage": round(coverage, 4),
            "unique_ratio": uniqueness,
            "duplicate_count": profile["duplicate_count"],
            "reason": "high uniqueness / keyword match",
        })

    candidates.extend(_rank_composite_key_candidates(column_profiles, data_rows, config))
    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates


def _rank_composite_key_candidates(
    column_profiles: list[dict[str, Any]],
    data_rows: int,
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    max_columns = int(config.get("auto_key_max_columns", 2))
    if max_columns < 2:
        return []

    usable = [
        profile for profile in column_profiles
        if profile.get("non_blank_count", 0) > 0
        and not _is_measure_like_column(profile, config)
    ]
    usable.sort(
        key=lambda p: (
            p.get("non_blank_count", 0) / max(data_rows, 1),
            p.get("unique_ratio", 0),
            _keyword_score(p.get("name", ""), config),
        ),
        reverse=True,
    )
    usable = usable[: int(config.get("auto_key_candidate_pool", 8))]

    result: list[dict[str, Any]] = []
    for i, left in enumerate(usable):
        for right in usable[i + 1:]:
            coverage = min(
                left.get("non_blank_count", 0) / max(data_rows, 1),
                right.get("non_blank_count", 0) / max(data_rows, 1),
            )
            estimated_unique = min(1.0, float(left.get("unique_ratio", 0)) * float(right.get("unique_ratio", 0)) * 1.1)
            if coverage < float(config.get("auto_key_min_composite_coverage", 0.05)):
                continue
            score = coverage * 35 + estimated_unique * 55 + max(
                _keyword_score(left.get("name", ""), config),
                _keyword_score(right.get("name", ""), config),
            ) - 2
            result.append({
                "columns": [left["name"], right["name"]],
                "score": round(score, 2),
                "coverage": round(coverage, 4),
                "unique_ratio": round(estimated_unique, 4),
                "duplicate_count": None,
                "reason": "composite key candidate from profile statistics",
            })
    return result


def _suggest_key_columns(
    base_sheet: dict[str, Any],
    rev_sheet: dict[str, Any],
    common_columns: list[str],
    config: dict[str, Any],
) -> list[str]:
    candidates: list[list[str]] = []
    profiles = _profiles_by_name(base_sheet, rev_sheet)
    for item in base_sheet.get("key_candidates", []) + rev_sheet.get("key_candidates", []):
        columns = [column for column in item.get("columns", []) if column in common_columns]
        if any(_is_measure_like_column(profiles.get(column, {}), config) for column in columns):
            continue
        if columns and len(columns) == len(item.get("columns", [])):
            candidates.append(columns)
    if not candidates:
        return []
    ranked = sorted(
        candidates,
        key=lambda columns: _key_quality(columns, base_sheet, rev_sheet),
        reverse=True,
    )
    return ranked[0]


def _key_quality(columns: list[str], base_sheet: dict[str, Any], rev_sheet: dict[str, Any]) -> tuple[float, float, float]:
    base_coverage, base_unique = _sample_key_quality(columns, base_sheet.get("row_samples", []) or [])
    rev_coverage, rev_unique = _sample_key_quality(columns, rev_sheet.get("row_samples", []) or [])
    coverage = min(base_coverage, rev_coverage)
    unique = min(base_unique, rev_unique)
    # Prefer short keys when quality is otherwise similar.
    return (unique, coverage, -len(columns))


def _sample_key_quality(columns: list[str], samples: list[dict[str, str]]) -> tuple[float, float]:
    if not columns or not samples:
        return 0.0, 0.0
    keys = []
    non_blank = 0
    for row in samples:
        parts = [str(row.get(column, "")).strip() for column in columns]
        if all(parts):
            non_blank += 1
            keys.append("|".join(parts))
    coverage = non_blank / max(len(samples), 1)
    unique = len(set(keys)) / max(len(keys), 1) if keys else 0.0
    return coverage, unique


def _suggest_structure_only_columns(base_sheet: dict[str, Any], rev_sheet: dict[str, Any]) -> list[str]:
    base_cols = set(base_sheet.get("columns", []))
    rev_cols = set(rev_sheet.get("columns", []))
    structure_only = []
    for column in sorted((base_cols | rev_cols) - (base_cols & rev_cols)):
        if _looks_auxiliary_column(column) or _is_generated_empty_column(column):
            structure_only.append(column)
    return structure_only


def _suggest_numeric_columns(base_sheet: dict[str, Any], rev_sheet: dict[str, Any], config: dict[str, Any]) -> dict[str, int]:
    numeric = {}
    for profile in _profiles_by_name(base_sheet, rev_sheet).values():
        if profile.get("numeric_ratio", 0) >= 0.8 and not _looks_id_column(profile["name"], config):
            numeric[profile["name"]] = 2
    return numeric


def _suggest_date_columns(base_sheet: dict[str, Any], rev_sheet: dict[str, Any], config: dict[str, Any]) -> list[str]:
    result = []
    for profile in _profiles_by_name(base_sheet, rev_sheet).values():
        if profile.get("date_like_ratio", 0) >= 0.6 and _looks_date_column(profile["name"], config):
            result.append(profile["name"])
    return result


def _profiles_by_name(base_sheet: dict[str, Any], rev_sheet: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result = {}
    for profile in base_sheet.get("column_profiles", []) + rev_sheet.get("column_profiles", []):
        current = result.get(profile["name"])
        if current is None or profile.get("non_blank_count", 0) > current.get("non_blank_count", 0):
            result[profile["name"]] = profile
    return result


def _choose_header_row(base_sheet: dict[str, Any], rev_sheet: dict[str, Any]) -> int:
    base = int(base_sheet.get("detected_header_row") or 1)
    rev = int(rev_sheet.get("detected_header_row") or 1)
    return base if base == rev else base


def _should_auto_include_sheet(base_sheet: dict[str, Any], rev_sheet: dict[str, Any], config: dict[str, Any]) -> bool:
    if config.get("include_sheets"):
        return True
    min_rows = int(config.get("auto_include_min_data_rows", 20))
    min_columns = int(config.get("auto_include_min_columns", 8))
    min_unique_ratio = float(config.get("auto_include_min_key_unique_ratio", 0.9))
    max_duplicate_header_ratio = float(config.get("auto_include_max_duplicate_header_ratio", 0.35))

    data_rows = max(base_sheet.get("data_rows", 0), rev_sheet.get("data_rows", 0))
    common_columns = set(base_sheet.get("columns", [])) & set(rev_sheet.get("columns", []))
    duplicate_header_ratio = max(
        float(base_sheet.get("duplicate_header_ratio", 0)),
        float(rev_sheet.get("duplicate_header_ratio", 0)),
    )
    key_candidates = [
        c for c in base_sheet.get("key_candidates", []) + rev_sheet.get("key_candidates", [])
        if c.get("columns", [None])[0] in common_columns
    ]
    best_unique_ratio = max(
        [c.get("unique_ratio", 0) for c in key_candidates]
        or [0]
    )

    return (
        data_rows >= min_rows
        and len(common_columns) >= min_columns
        and best_unique_ratio >= min_unique_ratio
        and duplicate_header_ratio <= max_duplicate_header_ratio
    )


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _looks_numeric(value: str) -> bool:
    try:
        float(str(value).replace(",", "").rstrip("%"))
        return True
    except ValueError:
        return False


def _looks_date_like(value: str) -> bool:
    return any(token in value for token in ("年", "月", "日", "-", "/", ".")) and any(ch.isdigit() for ch in value)


def _keyword_score(name: str, config: dict[str, Any]) -> int:
    lower = name.lower()
    scores = []
    for item in config.get("key_hint_keywords", []) or []:
        if isinstance(item, dict):
            keyword = str(item.get("keyword", ""))
            score = int(item.get("score", 6))
        else:
            keyword = str(item)
            score = 6
        if keyword and keyword.lower() in lower:
            scores.append(score)
    return max(scores, default=0)


def _looks_id_column(name: str, config: dict[str, Any]) -> bool:
    lowered = name.lower()
    return any(str(token).lower() in lowered for token in config.get("id_like_keywords", []) or [])


def _looks_date_column(name: str, config: dict[str, Any]) -> bool:
    lowered = name.lower()
    return any(str(token).lower() in lowered for token in config.get("date_like_keywords", []) or [])


def _is_measure_like_column(profile: dict[str, Any], config: dict[str, Any]) -> bool:
    name = str(profile.get("name", ""))
    lowered = name.lower()
    if any(str(token).lower() in lowered for token in config.get("key_exclude_keywords", []) or []):
        return True
    if profile.get("numeric_ratio", 0) > 0.95 and _keyword_score(name, config) < 6:
        return True
    return False


def _duplicate_header_ratio(columns: list[str]) -> float:
    if not columns:
        return 0.0
    base_names = [_dedupe_base_name(column) for column in columns]
    counts = Counter(base_names)
    duplicate_count = sum(count for count in counts.values() if count > 1)
    return round(duplicate_count / len(columns), 4)


def _dedupe_base_name(column: str) -> str:
    if "__" not in column:
        return column
    head, tail = column.rsplit("__", 1)
    return head if tail.isdigit() else column


def _looks_auxiliary_column(name: str) -> bool:
    lowered = name.strip().lower()
    if lowered.startswith(("tmp_", "aux_", "__")):
        return True
    if lowered.startswith("0") and any(ch.isalpha() or "\u4e00" <= ch <= "\u9fff" for ch in lowered):
        return True
    return False


def _is_generated_empty_column(name: str) -> bool:
    return name.startswith("__EMPTY_COL_")

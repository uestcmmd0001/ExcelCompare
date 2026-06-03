from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from compare_excel_to_excel.parsers.workbook_parser import parse_workbook
from compare_excel_to_excel.processing.differ import diff_workbooks


class ExcelDiffTests(unittest.TestCase):
    def test_business_key_prevents_row_insert_noise(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 10, 100],
                ["A002", "M2", 20, 200],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 10, 100],
                ["A999", "M9", 1, 9],
                ["A002", "M2", 20, 200],
            ])

            diffs = _run_diff(baseline, revised)
            self.assertEqual(_types(diffs).count("row_added"), 1)
            self.assertEqual(_types(diffs).count("cell_modified"), 0)
            self.assertIn("row_moved", _types(diffs))

    def test_added_deleted_and_modified_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 10, 100],
                ["A002", "M2", 20, 200],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 12, 100],
                ["A003", "M3", 30, 300],
            ])

            diffs = _run_diff(baseline, revised)
            self.assertIn("row_added", _types(diffs))
            self.assertIn("row_deleted", _types(diffs))
            self.assertIn("cell_modified", _types(diffs))
            qty_diff = [d for d in diffs if d.column_name == "metric_qty"][0]
            self.assertEqual(qty_diff.old_value, "10")
            self.assertEqual(qty_diff.new_value, "12")

    def test_ignored_column_and_numeric_precision(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value", "note"],
                ["A001", "M1", 100.001, "old"],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value", "note"],
                ["A001", "M1", 100.002, "new"],
            ])

            diffs = _run_diff(baseline, revised)
            self.assertEqual([d.diff_type for d in diffs if d.diff_type == "cell_modified"], [])

    def test_formula_diff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", "=10*2"],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", "=10*3"],
            ])

            diffs = _run_diff(baseline, revised)
            self.assertIn("formula_modified", _types(diffs))

    def test_sheet_and_column_structure_diff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value", "status"],
                ["A001", "M1", 100, "ok"],
            ], extra_sheet=True)

            diffs = _run_diff(baseline, revised)
            self.assertIn("column_added", _types(diffs))
            self.assertIn("sheet_added", _types(diffs))

    def test_auto_header_and_profile_suggest_key(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["公司抬头", None, None],
                [None, None, None],
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["公司抬头", None, None],
                [None, None, None],
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 120],
            ])

            from compare_excel_to_excel.config import load_config
            from compare_excel_to_excel.parsers.workbook_parser import parse_workbook
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            suggested = profile["suggested_config"]
            self.assertEqual(suggested["sheets"]["DataSheet"]["header_row"], 3)
            self.assertIn("record_id", suggested["sheets"]["DataSheet"]["key_columns"])

            baseline_wb = parse_workbook(baseline, suggested)
            revised_wb = parse_workbook(revised, suggested)
            diffs = diff_workbooks(baseline_wb, revised_wb, suggested)
            self.assertEqual(_types(diffs), ["cell_modified"])

    def test_effective_config_keeps_legacy_ignore_columns_compatible(self) -> None:
        from compare_excel_to_excel.main import _merge_config

        suggested = {
            "default": {"header_row": "auto", "key_columns": "auto"},
            "include_sheets": ["Sheet1"],
            "sheets": {
                "Sheet1": {
                    "header_row": 1,
                    "key_columns": ["record_id"],
                    "ignore_columns": ["aux_row_id"],
                }
            },
        }
        user_config = {
            "default": {"header_row": "auto", "key_columns": "auto"},
            "exclude_sheets": [],
            "sheets": {},
        }
        effective = _merge_config(suggested, user_config)
        self.assertEqual(effective["sheets"]["Sheet1"]["ignore_columns"], ["aux_row_id"])

    def test_column_added_is_structure_change_even_when_auxiliary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["0aux_source", "record_id", "item_id", "metric_value"],
                ["row-1", "A001", "M1", 100],
            ])

            config = {
                "default": {
                    "header_row": 1,
                    "key_columns": ["record_id", "item_id"],
                    "ignore_columns": ["0aux_source"],
                }
            }
            baseline_wb = parse_workbook(baseline, config)
            revised_wb = parse_workbook(revised, config)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(len(diffs), 1)
            self.assertEqual(diffs[0].diff_type, "column_added")
            self.assertEqual(diffs[0].result_layer, "structure_change")
            self.assertEqual(diffs[0].details["ignore_policy"], "ignored")
            self.assertTrue(diffs[0].need_human_review)

    def test_structure_only_column_added_is_reported_but_not_cell_compared(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["0aux_source", "record_id", "item_id", "metric_value"],
                ["row-1", "A001", "M1", 100],
            ])

            config = {
                "default": {
                    "header_row": 1,
                    "key_columns": ["record_id", "item_id"],
                    "structure_only_columns": ["0aux_source"],
                }
            }
            baseline_wb = parse_workbook(baseline, config)
            revised_wb = parse_workbook(revised, config)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(len(diffs), 1)
            self.assertEqual(diffs[0].diff_type, "column_added")
            self.assertEqual(diffs[0].result_layer, "structure_change")
            self.assertEqual(diffs[0].rule_source, "column_presence_structure_only")
            self.assertEqual(diffs[0].details["ignore_policy"], "structure_only")

    def test_duplicate_key_is_match_risk(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            rows = [
                ["record_id", "metric_desc", "metric_value"],
                ["A001", "first", 100],
                ["A001", "second", 200],
            ]
            _write_workbook(baseline, rows)
            _write_workbook(revised, rows)

            config = {"default": {"header_row": 1, "key_columns": ["record_id"]}}
            baseline_wb = parse_workbook(baseline, config)
            revised_wb = parse_workbook(revised, config)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(_types(diffs), ["duplicate_key", "duplicate_key_rows_skipped"])
            self.assertTrue(all(diff.result_layer == "match_risk" for diff in diffs))

    def test_duplicate_key_skips_confirmed_cell_diff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "metric_desc", "metric_value"],
                ["A001", "first", 100],
                ["A001", "second", 200],
            ])
            _write_workbook(revised, [
                ["record_id", "metric_desc", "metric_value"],
                ["A001", "first", 999],
                ["A001", "second", 200],
            ])

            config = {"default": {"header_row": 1, "key_columns": ["record_id"]}}
            baseline_wb = parse_workbook(baseline, config)
            revised_wb = parse_workbook(revised, config)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertIn("duplicate_key", _types(diffs))
            self.assertIn("duplicate_key_rows_skipped", _types(diffs))
            self.assertNotIn("cell_modified", _types(diffs))

    def test_fallback_plan_builds_valid_effective_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 10, 100],
                ["A002", "M2", 20, 200],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_qty", "metric_value"],
                ["A001", "M1", 11, 100],
                ["A002", "M2", 20, 200],
            ])

            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.planning.planner import build_comparison_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = build_comparison_plan(profile, config)
            validation = validate_plan(plan, profile, config)

            self.assertEqual(plan.primary_plan.sheet_pairs[0].baseline_sheet, "DataSheet")
            self.assertIn("DataSheet", validation.effective_config["include_sheets"])
            self.assertIn("record_id", validation.effective_config["sheets"]["DataSheet"]["key_columns"])
            self.assertGreater(validation.overall_confidence, 0.7)

    def test_fallback_plan_prefers_composite_key_when_single_key_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            rows = [
                ["record_id", "metric_desc", "metric_value"],
                ["A001", "first", 100],
                ["A001", "second", 200],
                ["A002", "first", 300],
            ]
            _write_workbook(baseline, rows)
            _write_workbook(revised, rows)

            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.planning.planner import build_comparison_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
                "planner_min_key_coverage": 0.5,
                "planner_min_key_uniqueness": 0.9,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = build_comparison_plan(profile, config)
            validation = validate_plan(plan, profile, config)
            key_columns = validation.effective_config["sheets"]["DataSheet"]["key_columns"]

            self.assertEqual(key_columns, ["record_id", "metric_desc"])
            baseline_wb = parse_workbook(baseline, validation.effective_config, role="baseline")
            revised_wb = parse_workbook(revised, validation.effective_config, role="revised")
            diffs = diff_workbooks(baseline_wb, revised_wb, validation.effective_config)
            self.assertEqual(diffs, [])

    def test_revised_header_offset_is_parsed_from_plan_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["title", None, None],
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 120],
            ])

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "include_sheets": ["DataSheet"],
                "sheets": {
                    "DataSheet": {
                        "header_row_baseline": 1,
                        "header_row_revised": 2,
                        "key_columns": ["record_id", "item_id"],
                    }
                },
            }
            baseline_wb = parse_workbook(baseline, config, role="baseline")
            revised_wb = parse_workbook(revised, config, role="revised")
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(_types(diffs), ["cell_modified"])
            self.assertEqual(diffs[0].row_before, 2)
            self.assertEqual(diffs[0].row_after, 3)

    def test_repeated_header_matrix_sheet_is_not_auto_included(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            repeated_columns = ["zone", "desc"] + ["same header"] * 10
            rows = [repeated_columns] + [[f"R{i}", f"Item {i}"] + [i] * 10 for i in range(30)]
            _write_workbook(baseline, rows)
            _write_workbook(revised, rows)

            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
                "auto_include_max_duplicate_header_ratio": 0.35,
            }
            profile = profile_workbook_pair(baseline, revised, config)

            self.assertEqual(profile["suggested_config"]["include_sheets"], [])
            self.assertGreater(profile["baseline"]["sheets"][0]["duplicate_header_ratio"], 0.35)

    def test_validator_normalizes_llm_sheet_and_column_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
                ["A002", "M2", 200],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 120],
                ["A002", "M2", 200],
            ])

            from compare_excel_to_excel.planning.plan_schema import ComparisonPlan, PlanCandidate, SheetPlan
            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = ComparisonPlan(
                version="1.0",
                planner="test_llm",
                source="llm",
                primary_plan=PlanCandidate(
                    candidate_id="test",
                    strategy="test",
                    confidence=0.8,
                    sheet_pairs=[
                        SheetPlan(
                            baseline_sheet="Data Sheet",
                            revised_sheet="Data Sheet",
                            header_row_baseline=1,
                            header_row_revised=1,
                            key_columns=["record id", "item id"],
                            column_mapping={
                                "record id": "record id",
                                "item id": "item id",
                                "metric value": "metric value",
                            },
                        )
                    ],
                ),
            )
            validation = validate_plan(plan, profile, config)

            self.assertIn("DataSheet", validation.effective_config["include_sheets"])
            self.assertEqual(
                validation.effective_config["sheets"]["DataSheet"]["key_columns"],
                ["record_id", "item_id"],
            )

    def test_validator_migrates_presence_only_ignore_columns_to_structure_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["0aux_source", "record_id", "item_id", "metric_value"],
                ["row-1", "A001", "M1", 100],
            ])

            from compare_excel_to_excel.planning.plan_schema import ComparisonPlan, PlanCandidate, SheetPlan
            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = ComparisonPlan(
                version="1.0",
                planner="test_llm",
                source="llm",
                primary_plan=PlanCandidate(
                    candidate_id="test",
                    strategy="test",
                    confidence=0.8,
                    sheet_pairs=[
                        SheetPlan(
                            baseline_sheet="DataSheet",
                            revised_sheet="DataSheet",
                            header_row_baseline=1,
                            header_row_revised=1,
                            key_columns=["record_id", "item_id"],
                            column_mapping={
                                "record_id": "record_id",
                                "item_id": "item_id",
                                "metric_value": "metric_value",
                            },
                            ignore_columns=["0 aux source"],
                        )
                    ],
                ),
            )
            validation = validate_plan(plan, profile, config)
            sheet_cfg = validation.effective_config["sheets"]["DataSheet"]

            self.assertEqual(sheet_cfg["ignore_columns"], [])
            self.assertEqual(sheet_cfg["structure_only_columns"], ["0aux_source"])

    def test_validator_skips_matrix_sheet_from_executable_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            repeated_columns = ["zone", "desc"] + ["same header"] * 10
            rows = [repeated_columns] + [[f"R{i}", f"Item {i}"] + [i] * 10 for i in range(30)]
            _write_workbook(baseline, rows)
            _write_workbook(revised, rows)

            from compare_excel_to_excel.planning.plan_schema import ComparisonPlan, PlanCandidate, SheetPlan
            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "include_sheets": ["DataSheet"],
                "auto_include_max_duplicate_header_ratio": 0.35,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = ComparisonPlan(
                version="1.0",
                planner="test_llm",
                source="llm",
                primary_plan=PlanCandidate(
                    candidate_id="test",
                    strategy="test",
                    confidence=0.8,
                    sheet_pairs=[
                        SheetPlan(
                            baseline_sheet="DataSheet",
                            revised_sheet="DataSheet",
                            header_row_baseline=1,
                            header_row_revised=1,
                            key_columns=["zone"],
                            column_mapping={column: column for column in profile["baseline"]["sheets"][0]["columns"]},
                        )
                    ],
                ),
            )
            validation = validate_plan(plan, profile, config)

            self.assertEqual(validation.effective_config["include_sheets"], [])
            self.assertIn("skipped from executable diff", " ".join(validation.risks))

    def test_skipped_executable_sheets_do_not_fall_back_to_deep_diff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 999],
            ])

            config = {
                "default": {"header_row": 1, "key_columns": ["record_id", "item_id"]},
                "include_sheets": [],
            }
            baseline_wb = parse_workbook(baseline, config, include_all_sheets=True)
            revised_wb = parse_workbook(revised, config, include_all_sheets=True)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(diffs, [])

    def test_sheet_added_deleted_are_reported_outside_executable_sheet_filter(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ], extra_sheet=True)
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _add_sheet(revised, "新增表")

            config = {
                "default": {"header_row": 1, "key_columns": ["record_id", "item_id"]},
                "include_sheets": ["DataSheet"],
            }
            baseline_wb = parse_workbook(baseline, config, include_all_sheets=True)
            revised_wb = parse_workbook(revised, config, include_all_sheets=True)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertIn("sheet_added", _types(diffs))
            self.assertIn("sheet_deleted", _types(diffs))

    def test_column_mapping_compares_renamed_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "old_metric"],
                ["A001", "M1", 100],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "new_metric"],
                ["A001", "M1", 120],
            ])

            config = {
                "default": {"header_row": 1, "key_columns": ["record_id", "item_id"]},
                "sheets": {
                    "DataSheet": {
                        "column_mapping": {
                            "record_id": "record_id",
                            "item_id": "item_id",
                            "old_metric": "new_metric",
                        }
                    }
                },
            }
            baseline_wb = parse_workbook(baseline, config)
            revised_wb = parse_workbook(revised, config)
            diffs = diff_workbooks(baseline_wb, revised_wb, config)

            self.assertEqual(_types(diffs), ["cell_modified"])
            self.assertEqual(diffs[0].column_name, "old_metric -> new_metric")
            self.assertEqual(diffs[0].old_value, "100")
            self.assertEqual(diffs[0].new_value, "120")

    def test_cross_name_sheet_pair_is_executable(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_named_workbook(baseline, "OldSheet", [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 100],
            ])
            _write_named_workbook(revised, "NewSheet", [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", 120],
            ])

            from compare_excel_to_excel.planning.plan_schema import ComparisonPlan, PlanCandidate, SheetPlan
            from compare_excel_to_excel.planning.plan_validator import validate_plan
            from compare_excel_to_excel.profilers.workbook_profiler import profile_workbook_pair

            config = {
                "default": {"header_row": "auto", "key_columns": "auto"},
                "auto_include_min_data_rows": 1,
                "auto_include_min_columns": 2,
                "auto_include_min_key_unique_ratio": 0.5,
            }
            profile = profile_workbook_pair(baseline, revised, config)
            plan = ComparisonPlan(
                version="1.0",
                planner="test_llm",
                source="llm",
                primary_plan=PlanCandidate(
                    candidate_id="test",
                    strategy="sheet_alias",
                    confidence=0.8,
                    sheet_pairs=[
                        SheetPlan(
                            baseline_sheet="OldSheet",
                            revised_sheet="NewSheet",
                            header_row_baseline=1,
                            header_row_revised=1,
                            key_columns=["record_id", "item_id"],
                            column_mapping={
                                "record_id": "record_id",
                                "item_id": "item_id",
                                "metric_value": "metric_value",
                            },
                        )
                    ],
                ),
            )
            validation = validate_plan(plan, profile, config)

            self.assertEqual(
                validation.effective_config["sheet_pairs"],
                [{"baseline_sheet": "OldSheet", "revised_sheet": "NewSheet"}],
            )

            baseline_wb = parse_workbook(baseline, validation.effective_config, role="baseline", include_all_sheets=True)
            revised_wb = parse_workbook(revised, validation.effective_config, role="revised", include_all_sheets=True)
            diffs = diff_workbooks(baseline_wb, revised_wb, validation.effective_config)

            self.assertEqual(_types(diffs), ["cell_modified"])
            self.assertEqual(diffs[0].sheet_name, "OldSheet -> NewSheet")
            self.assertEqual(diffs[0].details["baseline_sheet"], "OldSheet")
            self.assertEqual(diffs[0].details["revised_sheet"], "NewSheet")
            self.assertEqual(diffs[0].details["canonical_column"], "metric_value")

    def test_formula_text_change_without_cached_value_is_not_double_counted(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            baseline = Path(tmp) / "baseline.xlsx"
            revised = Path(tmp) / "revised.xlsx"

            _write_workbook(baseline, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", "=10*2"],
            ])
            _write_workbook(revised, [
                ["record_id", "item_id", "metric_value"],
                ["A001", "M1", "=10*3"],
            ])

            diffs = _run_diff(baseline, revised)

            self.assertEqual(_types(diffs), ["formula_modified"])


def _run_diff(baseline: Path, revised: Path):
    config = {
        "default": {
            "header_row": 1,
            "key_columns": ["record_id", "item_id"],
            "ignore_columns": ["note"],
            "numeric_columns": {"metric_value": 2, "metric_qty": 0},
            "compare_formulas": True,
        }
    }
    baseline_wb = parse_workbook(baseline, config)
    revised_wb = parse_workbook(revised, config)
    return diff_workbooks(baseline_wb, revised_wb, config)


def _write_workbook(path: Path, rows: list[list], extra_sheet: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    for row in rows:
        ws.append(row)
    if extra_sheet:
        extra = wb.create_sheet("附表")
        extra.append(["字段"])
        extra.append(["值"])
    wb.save(path)


def _write_named_workbook(path: Path, sheet_name: str, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


def _add_sheet(path: Path, title: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.create_sheet(title)
    ws.append(["record_id", "item_id", "metric_value"])
    ws.append(["A999", "M9", 9])
    wb.save(path)


def _types(diffs) -> list[str]:
    return [d.diff_type for d in diffs]


if __name__ == "__main__":
    unittest.main()
